# -*- coding: utf-8 -*-
"""
Created on Wed Nov 23 10:44:51 2022

@author: WUPEIZH
"""

# -*- coding: utf-8 -*-
"""
Created on Mon May 16 16:52:41 2022

@author: WUPEIZH

This code is to export a z-stack with image translated and rotated using affine transformation 
This code also export SG.V6, OC.V6, RC.V6 with the coordinates corrected by the transformation 
"""

#%% Import packages
import slideio
import cv2
import os
import math
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from tifffile import imwrite
import matplotlib.pyplot as plt
import re
import matplotlib.colors as mcolors
import matplotlib.cm as cmx
import fnmatch
#%% Define functions for 3D recon
def savecor(listofcor,remove=[],scale = 1,):
    Arrayoflist = np.asarray(listofcor)
    Arrayoflist= Arrayoflist*scale
    Arrayoflist = np.delete(Arrayoflist, remove)
    tofile = np.array(Arrayoflist).T.tolist()
    return tofile

def rescale(listofcor,scale = 1,Firstslide=0):
    Arrayoflist = np.asarray(listofcor)
    Arrayoflist= (Arrayoflist-Firstslide)*scale
    Arrayoflist= Arrayoflist.astype(np.float)
    return Arrayoflist

def Readcoordinate(datasource,column,coordinate,readfromrow):
    thecolumn = datasource[column]
    for cell in thecolumn[readfromrow:]:
        coordinate = coordinate+[cell.value]
    return coordinate    


         
#%% Define functions for reading files
def find_crop_coordinate(image,cropwidth = 2400,cropheight=3600):
    cropwidth = int(cropwidth)
    cropheight = int(cropheight)
    [imageheight,imagewidth,c]=image.shape
    find_crop_coordinate.NX = math.floor(imagewidth/cropwidth)
    find_crop_coordinate.NY = math.floor(imageheight/cropheight)
    #print("number of x tile",find_crop_coordinate.NX, '\n', "number of y tile",find_crop_coordinate.NY)    

def readexcel(ReadfromDatasource,rows):
    readexcel.Filelist = [ReadfromDatasource.cell(row = rows,column = 2).value]
    readexcel.X_cor = [ReadfromDatasource.cell(row = rows,column = 3).value]
    readexcel.Y_cor = [ReadfromDatasource.cell(row = rows,column = 4).value]
    readexcel.Width_cor =  [ReadfromDatasource.cell(row = rows,column = 5).value]
    readexcel.Height_cor =  [ReadfromDatasource.cell(row = rows,column = 6).value]
    readexcel.folder = [ReadfromDatasource.cell(row = rows,column = 8).value]
    readexcel.Tray = [ReadfromDatasource.cell(row = rows,column = 1).value]


#%% define functions for image processing
def read_slide(full,xstart = 40443,ystart =16254,width = 4995,height = 3333,scale=1):
    slide = slideio.open_slide(path=full,driver = 'SVS')
    scene = slide.get_scene(0)
    read_slide.image = scene.read_block(rect=(xstart,ystart,width,height),size = (math.ceil(width/scale),math.ceil(height/scale)),slices=(0,0))
    print("Image imported from "+full)

def crop_image(image,cropwidth=2400,cropheight=3600,cropxstart = 0,cropystart = 0,name = '878.SVS',roundn = 0,path = r'C:\Users\WUPEIZH\Desktop'):
    cropped = image[cropystart:cropystart+cropheight,cropxstart:cropxstart+cropwidth]
    #plt.imshow(cropped)    
    background = np.zeros([3600,2400,3]);
    background = background.astype(np.uint8)
    background[0:cropped.shape[0],0:cropped.shape[1]]=cropped
    background_after = np.rot90(background)
    background_after = cv2.cvtColor(background_after,cv2.COLOR_BGR2RGB)
    name = 'segment'+ str(roundn)+'.tif'
    cv2.imwrite(os.path.join(path,name),background_after)
    print("Image written for " + path+name)

          
#%% Define functions for reading files
def find_crop_coordinate(image,cropwidth = 2400,cropheight=3600):
    cropwidth = int(cropwidth)
    cropheight = int(cropheight)
    [imageheight,imagewidth,c]=image.shape
    find_crop_coordinate.NX = math.floor(imagewidth/cropwidth)
    find_crop_coordinate.NY = math.floor(imageheight/cropheight)
    #print("number of x tile",find_crop_coordinate.NX, '\n', "number of y tile",find_crop_coordinate.NY)    

def readexcel(ReadfromDatasource,rows):
    readexcel.Filelist = [ReadfromDatasource.cell(row = rows,column = 2).value]
    readexcel.X_cor = [ReadfromDatasource.cell(row = rows,column = 3).value]
    readexcel.Y_cor = [ReadfromDatasource.cell(row = rows,column = 4).value]
    readexcel.Width_cor =  [ReadfromDatasource.cell(row = rows,column = 5).value]
    readexcel.Height_cor =  [ReadfromDatasource.cell(row = rows,column = 6).value]
    readexcel.folder = [ReadfromDatasource.cell(row = rows,column = 8).value]
    readexcel.Tray = [ReadfromDatasource.cell(row = rows,column = 1).value]
    readexcel.Rotationalangle = [ReadfromDatasource.cell(row = rows,column = 9).value]
    readexcel.XcenterOri = [ReadfromDatasource.cell(row = rows,column = 10).value]
    readexcel.YcenterOri = [ReadfromDatasource.cell(row = rows,column = 11).value]
    readexcel.Xcentertranslated = [ReadfromDatasource.cell(row = rows,column = 12).value]
    readexcel.Ycentertranslated = [ReadfromDatasource.cell(row = rows,column = 13).value]
    
    
#%% Define functions to crop and rotate image 
def cropcochlearcenter(readexcel,x_HC_Cor,y_HC_Cor,z_Slide):
    searchslide = readexcel.Filelist[0]
    searchslide = re.search(Side+'(.*).svs',searchslide)
    newslide = int(searchslide.group(1))
    indices = [i for i, x in enumerate(z_Slide) if x == newslide]
    crop_x,crop_y,crop_width,crop_height,x_center,y_center,topleft,rectwidth=[],[],[],[],[],[],[],[]
    if indices:
        x_cor,y_cor = [],[]
        for i in indices:
            x_cor,y_cor = x_cor+[x_HC_Cor[i]],y_cor+[y_HC_Cor[i]]
        x_center = (max(x_cor)+min(x_cor))/2
        y_center = (max(y_cor)+min(y_cor))/2      
        width = max(x_cor)-min(x_cor)
        height = max(y_cor)-min(y_cor)
        rectwidth = max([20000,int(math.ceil((width+height)/2))])
        topleft = int(x_center-math.ceil(rectwidth/2)),int(y_center-math.ceil(rectwidth/2))
 
        # x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
        # boxwidth = min(x_center,y_center,width_T-x_center,height_T-y_center,4000)
        # crop_x,crop_y,crop_width,crop_height = [int(x_cor_T+x_center-boxwidth),int(y_cor_T+y_center-boxwidth),int(boxwidth*2),int(boxwidth*2)]
    return topleft,x_center,y_center,rectwidth

def imagerotation(originalslide,rotationalangle = 0):
    height, width = originalslide.shape[:2]
    centerx,centery = (width/2, height/2)
    rotation_mat = cv2.getRotationMatrix2D((centerx,centery),rotationalangle,1)
    abs_cos = abs(rotation_mat[0,0])
    abs_sin = abs(rotation_mat[0,1])
    bound_w = int(height * abs_sin + width * abs_cos)
    bound_h = int(height * abs_cos + width * abs_sin)
    rotation_mat[0, 2] += bound_w/2 - centerx
    rotation_mat[1, 2] += bound_h/2 - centery
    rotatedSlide2 = cv2.warpAffine(originalslide,rotation_mat,(bound_w,bound_h))
    background = rotatedSlide2.copy()
    rotatedH,rotatedW = background.shape[:2]
    centerx,centery = (math.ceil(rotatedW/2),math.ceil(rotatedH/2))
    croppedbackground = background[centerx-math.ceil(rotatedH/2.2):centerx+math.ceil(rotatedH/2.2),centery-math.ceil(rotatedH/2.2):centery+math.ceil(rotatedH/2.2)]
    croppedbackground = background[centerx-math.ceil(height*0.4):centerx+math.ceil(height*0.4),centery-math.ceil(height*0.4):centery+math.ceil(height*0.4)]
    return croppedbackground

#%% functions to assign angle to spreadsheets
def transferrotationcenter(excel,rotationalangle,rotationalcenterx,rotationalcentery,wholestackslidenumber,rotationalcenterx_shift,rotationalcentery_shift):
    Readactivated = excel.active
    slidenumber = Readcoordinate(Readactivated,'D',[],1)
    for n,slide in enumerate(slidenumber):
        Readactivated.cell(row = n+2,column = 12).value = rotationalangle[wholestackslidenumber.index(str(slide))]
        Readactivated.cell(row = n+2,column = 13).value = rotationalcenterx[wholestackslidenumber.index(str(slide))]
        Readactivated.cell(row = n+2,column = 14).value = rotationalcentery[wholestackslidenumber.index(str(slide))]
        Readactivated.cell(row = n+2,column = 15).value = rotationalcenterx_shift[wholestackslidenumber.index(str(slide))]
        Readactivated.cell(row = n+2,column = 16).value = rotationalcentery_shift[wholestackslidenumber.index(str(slide))]
        
    Readactivated.cell(row = 1,column = 12).value = 'rotation angle centered around center of organ of corti'
    Readactivated.cell(row = 1,column = 13).value = 'rotation x center'
    Readactivated.cell(row = 1,column = 14).value = 'rotation y center'
    Readactivated.cell(row = 1,column = 15).value = 'shifted rotation x center'
    Readactivated.cell(row = 1,column = 16).value = 'shifted rotation y center'
    
    return Readactivated


def loadcoordinate(HCDatasource):
    ReadfromDatasource = HCDatasource.active
    rotateangle,rotateXcenter,rotateYcenter,rotateXcenter_shifted,rotateYcenter_shifted,x_cor,y_cor,z_slide = [Readcoordinate(ReadfromDatasource,'L',[],1),
                                    Readcoordinate(ReadfromDatasource,'M',[],1),
                                    Readcoordinate(ReadfromDatasource,'N',[],1),
                                    Readcoordinate(ReadfromDatasource,'O',[],1),
                                    Readcoordinate(ReadfromDatasource,'P',[],1),
                                    Readcoordinate(ReadfromDatasource,'B',[],1), 
                                    Readcoordinate(ReadfromDatasource,'C',[],1), 
                                    Readcoordinate(ReadfromDatasource,'D',[],1)]
    Allcoordinate =[x_cor,y_cor,rotateangle,rotateXcenter,rotateYcenter,rotateXcenter_shifted,rotateYcenter_shifted]
    Allcoordinate = np.array([Allcoordinate]).T.tolist()  
    return Allcoordinate,z_slide


def rotationnormalization(Allcoordinate,z_slide,Allcoordinate_HC,z_slide_HC):
    if not Allcoordinate_HC :
        Allcoordinate_HC,z_slide_HC = Allcoordinate,z_slide
    sortedslide = sorted(set(z_slide_HC))
    for slide in sortedslide[:-1]:
        index = []
        for x in enumerate(z_slide):
            if x[1]<=int(slide):
                index = index+[x[0]]
        index_current = z_slide_HC.index(int(slide))
        rotateangle,rotateXcenter,rotateYcenter,rotateXcenter_shifted,rotateYcenter_shifted = Allcoordinate_HC[index_current][2:]
        for ind in index: 
            x_ORI = Allcoordinate[ind][0]
            y_ORI = Allcoordinate[ind][1]
            corrected_x_ro,corrected_y_ro =  rotate((rotateXcenter[0],rotateYcenter[0]),(x_ORI[0],y_ORI[0]),math.radians(-rotateangle[0]))
            corrected_x = corrected_x_ro+(rotateXcenter_shifted[0]-rotateXcenter[0])
            corrected_y = corrected_y_ro+(rotateYcenter_shifted[0]-rotateYcenter[0])
            Allcoordinate[ind][0] = [corrected_x]
            Allcoordinate[ind][1] = [corrected_y]
    
    corrected_x_cor,corrected_y_cor=[],[]
    for n,slide in enumerate(z_slide):
        corrected_x_cor = corrected_x_cor+ Allcoordinate[n][0]
        corrected_y_cor = corrected_y_cor+ Allcoordinate[n][1]                
    x_normalized,y_normalized= rescale(corrected_x_cor,scale), rescale(corrected_y_cor,scale)
    return x_normalized,y_normalized

def enternormalizeddatabacktospreasheet(Datasource,x_normalized,y_normalized ):
    for n,x in enumerate(x_normalized):
       activated = Datasource.active
       activated.cell(row = n+2,column = 17).value = x
       activated.cell(row = n+2,column = 18).value = y_normalized[n]
    activated.cell(row = 1,column = 17).value = 'Normalized X'
    activated.cell(row = 1,column = 18).value ='Normalized Y'
    return activated



#%% define helper function 
def find(pattern, path):
    result = []
    for root, dirs, files in os.walk(path):

        for name in files:
            if fnmatch.fnmatch(name, pattern):
                result.append(os.path.join(root,name))
    return result

def calculateaffinetransformation(slidenumber,HC_slide,HC_local_X_Cor,HC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_HC,Transformed_Y_cor_HC):
    HClocationforthisslide = int(slidenumber) in HC_slide
    if HClocationforthisslide:
        # plt.figure()
        # plt.imshow(im2_gray)
        for k,allslide in enumerate(HC_slide):
            if allslide == int(slidenumber):
                HCloc = [[HC_local_X_Cor[k],HC_local_Y_Cor[k]]]
                
                # plt.imshow(im2_gray)
                box = np.array([HCloc])/zoomout
                # plt.scatter(box[0][0][0],box[0][0][1])
                
                rotatedbox = cv2.transform(box,warp_matrix) #*zoomout
                rotatedbox[0][0][1] = 2*box[0][0][1]-(rotatedbox[0][0][1])
                rotatedbox[0][0][0] = 2*box[0][0][0]-(rotatedbox[0][0][0])
                # plt.imshow(im2_aligned)
                # plt.scatter(rotatedbox[0][0][0],rotatedbox[0][0][1])
                
                
                Transformed_X_cor_HC[k] = rotatedbox[0][0][0]*zoomout
                Transformed_Y_cor_HC[k] = rotatedbox[0][0][1]*zoomout
    return(Transformed_X_cor_HC,Transformed_Y_cor_HC)

#%%  The alignement code 

Row_INDEX = Row_index = [58,104]
rerun20x = []
cannotconverge = []
success = []
for Row_index in Row_INDEX:
    try:
        Indextoaddress  = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Code and data  for figures v10\Index to addressbook.xlsx')
        ReadIndextoaddress = Indextoaddress.active
        traynumber = ReadIndextoaddress.cell(Row_index,column = 1).value
        Side = ReadIndextoaddress.cell(Row_index,column = 2).value
        print('working on ',str(traynumber))
        firstslide = ReadIndextoaddress.cell(Row_index,column = 6).value
        first_line = ReadIndextoaddress.cell(Row_index,column = 3).value # the first line in excel to read from 
        MidModiolar_line = ReadIndextoaddress.cell(Row_index,column = 4).value
        last_line = ReadIndextoaddress.cell(Row_index,column = 5).value # the last line in excel to read
        scale = 1/1994
        zoomout = 20
        Datasource = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Code and data  for figures v10\Address book for cases in my database.xlsx')
        ReadfromDatasource = Datasource.active
        rootdatapathformask = r'\\apollo\research\ENT\Liberman\Peizhe Wu\01 Human study WPZ\01 Human Ear Raw Data\Human temporal bone 3D stack'
        datapathformask = os.path.join(rootdatapathformask,'Tray'+str(traynumber),'Mask')
        datapathforcoordinate = os.path.join(rootdatapathformask,'Tray'+str(traynumber))
        localpath = os.path.join('D:\SGC project data\Human temporal bone 3D stack','Tray'+str(traynumber))
        newdatapathformask= os.path.join(rootdatapathformask,'Tray'+str(traynumber),'Transformed Mask')
        Allmaskfiles = find('*merge_mask.tif', datapathformask)
        if os.path.isfile(os.path.join(datapathforcoordinate,'xyz coordinates for HC.xlsx')):
    
            HCdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for HC.xlsx'))
            HCdatafromdatasource = HCdatasource.active
            HC_slide,HC_local_X_Cor,HC_local_Y_Cor,HC_local_Z_Cor = [Readcoordinate(HCdatafromdatasource,'D',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'B',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'C',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'G',[],1)]
            Transformed_X_cor_HC,Transformed_Y_cor_HC = HC_local_X_Cor.copy(),HC_local_Y_Cor.copy()
          
                        
                        
            RCdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for RC.xlsx'))
            RCdatafromdatasource = RCdatasource.active
            RC_slide,RC_local_X_Cor,RC_local_Y_Cor,RC_local_Z_Cor = [Readcoordinate(RCdatafromdatasource,'D',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'B',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'C',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'G',[],1)]
            Transformed_X_cor_RC,Transformed_Y_cor_RC = RC_local_X_Cor.copy(),RC_local_Y_Cor.copy()
            
            SGNdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for SGN.xlsx'))
            SGNdatafromdatasource = SGNdatasource.active
            SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,SGN_local_Z_Cor = [Readcoordinate(SGNdatafromdatasource,'D',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'B',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'C',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'G',[],1)]
            Transformed_X_cor_SGN,Transformed_Y_cor_SGN = SGN_local_X_Cor.copy(),SGN_local_Y_Cor.copy()
            
            # Implantdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for implant.xlsx'))
            # Implantdatafromdatasource = Implantdatasource.active
            # Implant_slide,Implant_local_X_Cor,Implant_local_Y_Cor,Implant_local_Z_Cor = [Readcoordinate(SGNdatafromdatasource,'D',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'B',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'C',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'G',[],1)]
            # Transformed_X_cor_Implant,Transformed_Y_cor_Implant = Implant_local_X_Cor.copy(),Implant_local_Y_Cor.copy()
        
            for n,row in enumerate(np.arange(MidModiolar_line,last_line,1).astype(int)):    
            # for n,row in enumerate(np.arange(3520,last_line,1).astype(int)):
                direction = 1
                print(row)
                readexcel(ReadfromDatasource,row)
                if n == 0:
                    x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                    Slide1 = read_slide.image
                    stack =np.zeros([50,int(math.ceil(height_T/zoomout)),int(math.ceil(width_T/zoomout)),3],dtype=np.uint8)
                else:    
                    Slide1 = im2_aligned
                readexcel(ReadfromDatasource,row+1)
                x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
        
                read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                Slide2 = read_slide.image
                
                im1_gray = cv2.cvtColor(Slide1,cv2.COLOR_BGR2GRAY)
                im2_gray = cv2.cvtColor(Slide2,cv2.COLOR_BGR2GRAY)
                sz = im1_gray.shape
                warp_mode = cv2.MOTION_EUCLIDEAN 
                warp_matrix = np.eye(2, 3, dtype=np.float32)
                number_of_iterations = 5000;
                termination_eps = 1e-10;
                criteria = (cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, number_of_iterations,  termination_eps)
                (cc, warp_matrix) = cv2.findTransformECC(im1_gray,im2_gray,warp_matrix, warp_mode, criteria)
                im2_aligned = cv2.warpAffine(Slide2, warp_matrix, (sz[1],sz[0]), flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
                stack[row-first_line,0:sz[0],0:sz[1],:] = im2_aligned
                Sli = re.search(Side+'(.*).svs',readexcel.Filelist[0])
                if 'r' in Sli.group(1):
                    index = Sli.group(1).index('r') 
                    slidenumber = Sli.group(1)[0:index]
                else:
                    slidenumber = Sli.group(1)
                
                Transformed_X_cor_HC,Transformed_Y_cor_HC = calculateaffinetransformation(slidenumber,HC_slide,HC_local_X_Cor,HC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_HC,Transformed_Y_cor_HC)
                Transformed_X_cor_SGN,Transformed_Y_cor_SGN = calculateaffinetransformation(slidenumber,SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_SGN,Transformed_Y_cor_SGN)
                Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
                # Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
        
            for n,row in enumerate(np.arange(MidModiolar_line+1,first_line+1,-1).astype(int)):
                direction = -1
                print(row)
                readexcel(ReadfromDatasource,row)
                if n == 0:
                    x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                    Slide1 = read_slide.image
                else:    
                    Slide1 = im2_aligned
                readexcel(ReadfromDatasource,row-1)
                x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                Slide2 = read_slide.image
                im1_gray = cv2.cvtColor(Slide1,cv2.COLOR_BGR2GRAY)
                im2_gray = cv2.cvtColor(Slide2,cv2.COLOR_BGR2GRAY)
                sz = im1_gray.shape
                warp_mode = cv2.MOTION_EUCLIDEAN 
                warp_matrix = np.eye(2, 3, dtype=np.float32)
                number_of_iterations = 5000;
                termination_eps = 1e-10;
                criteria = (cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, number_of_iterations,  termination_eps)
                (cc, warp_matrix) = cv2.findTransformECC (im1_gray,im2_gray,warp_matrix, warp_mode, criteria)
                im2_aligned = cv2.warpAffine(Slide2, warp_matrix, (sz[1],sz[0]), flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
                
                stack[row-first_line,0:sz[0],0:sz[1],:] = im2_aligned
                Sli = re.search(Side+'(.*).svs',readexcel.Filelist[0])
                if 'r' in Sli.group(1):
                    index = Sli.group(1).index('r') #number+'r' is rescanned image
                    slidenumber = Sli.group(1)[0:index]
                else:
                    slidenumber = Sli.group(1)
                Transformed_X_cor_HC,Transformed_Y_cor_HC = calculateaffinetransformation(slidenumber,HC_slide,HC_local_X_Cor,HC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_HC,Transformed_Y_cor_HC)
                Transformed_X_cor_SGN,Transformed_Y_cor_SGN = calculateaffinetransformation(slidenumber,SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_SGN,Transformed_Y_cor_SGN)
                Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
            To_list =np.array( [Transformed_X_cor_HC,Transformed_Y_cor_HC,HC_slide,HC_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for HC v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for HC v6.xlsx'))  
            
            To_list =np.array( [Transformed_X_cor_RC,Transformed_Y_cor_RC,RC_slide,RC_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for RC v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for RC v6.xlsx'))  
            
            To_list =np.array( [Transformed_X_cor_SGN,Transformed_Y_cor_SGN,SGN_slide,SGN_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for SGN v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for SGN v6.xlsx'))
            success = success + [Row_index]
    except:
        rerun20x = rerun20x+[Row_index]
        print(' error')
        print(rerun20x)
        pass

    
#%% 
rerun20x = rerun20x+[38,41,42,46,48,52,66,53,57,58,64,65,67]
cannotconverge = []
successin20x = []
Row_INDEX = rerun20x
for Row_index in Row_INDEX:
    try:
        Indextoaddress  = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Code and data  for figures v10\Index to addressbook.xlsx')
        ReadIndextoaddress = Indextoaddress.active
        traynumber = ReadIndextoaddress.cell(Row_index,column = 1).value
        Side = ReadIndextoaddress.cell(Row_index,column = 2).value
        print('working on ',str(traynumber))
        firstslide = ReadIndextoaddress.cell(Row_index,column = 6).value
        first_line = ReadIndextoaddress.cell(Row_index,column = 3).value # the first line in excel to read from 
        MidModiolar_line = ReadIndextoaddress.cell(Row_index,column = 4).value
        last_line = ReadIndextoaddress.cell(Row_index,column = 5).value # the last line in excel to read
        scale = 1/1994
        zoomout = 20
        Datasource = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Code and data  for figures v10\Address book for cases in my database.xlsx')
        ReadfromDatasource = Datasource.active
        rootdatapathformask = r'\\apollo\research\ENT\Liberman\Peizhe Wu\01 Human study WPZ\01 Human Ear Raw Data\Human temporal bone 3D stack'
        datapathformask = os.path.join(rootdatapathformask,'Tray'+str(traynumber),'Mask')
        datapathforcoordinate = os.path.join(rootdatapathformask,'Tray'+str(traynumber))
        localpath = os.path.join('D:\SGC project data\Human temporal bone 3D stack','Tray'+str(traynumber))
        newdatapathformask= os.path.join(rootdatapathformask,'Tray'+str(traynumber),'Transformed Mask')
        Allmaskfiles = find('*merge_mask.tif', datapathformask)
        if os.path.isfile(os.path.join(datapathforcoordinate,'xyz coordinates for HC v4.xlsx')):
    
            HCdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for HC.xlsx'))
            HCdatafromdatasource = HCdatasource.active
            HC_slide,HC_local_X_Cor,HC_local_Y_Cor,HC_local_Z_Cor = [Readcoordinate(HCdatafromdatasource,'D',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'B',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'C',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'G',[],1)]
            Transformed_X_cor_HC,Transformed_Y_cor_HC = HC_local_X_Cor.copy(),HC_local_Y_Cor.copy()
          
                        
                        
            RCdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for RC.xlsx'))
            RCdatafromdatasource = RCdatasource.active
            RC_slide,RC_local_X_Cor,RC_local_Y_Cor,RC_local_Z_Cor = [Readcoordinate(RCdatafromdatasource,'D',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'B',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'C',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'G',[],1)]
            Transformed_X_cor_RC,Transformed_Y_cor_RC = RC_local_X_Cor.copy(),RC_local_Y_Cor.copy()
            
            SGNdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for SGN.xlsx'))
            SGNdatafromdatasource = SGNdatasource.active
            SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,SGN_local_Z_Cor = [Readcoordinate(SGNdatafromdatasource,'D',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'B',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'C',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'G',[],1)]
            Transformed_X_cor_SGN,Transformed_Y_cor_SGN = SGN_local_X_Cor.copy(),SGN_local_Y_Cor.copy()
            
            # Implantdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for implant.xlsx'))
            # Implantdatafromdatasource = Implantdatasource.active
            # Implant_slide,Implant_local_X_Cor,Implant_local_Y_Cor,Implant_local_Z_Cor = [Readcoordinate(SGNdatafromdatasource,'D',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'B',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'C',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'G',[],1)]
            # Transformed_X_cor_Implant,Transformed_Y_cor_Implant = Implant_local_X_Cor.copy(),Implant_local_Y_Cor.copy()
        
            for n,row in enumerate(np.arange(MidModiolar_line,last_line,1).astype(int)):    
            # for n,row in enumerate(np.arange(3520,last_line,1).astype(int)):
                direction = 1
                print(row)
                readexcel(ReadfromDatasource,row)
                if n == 0:
                    x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                    Slide1 = read_slide.image
                    stack =np.zeros([50,int(math.ceil(height_T/zoomout)),int(math.ceil(width_T/zoomout)),3],dtype=np.uint8)
                else:    
                    Slide1 = im2_aligned
                readexcel(ReadfromDatasource,row+1)
                x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
        
                read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                Slide2 = read_slide.image
                
                im1_gray = cv2.cvtColor(Slide1,cv2.COLOR_BGR2GRAY)
                im2_gray = cv2.cvtColor(Slide2,cv2.COLOR_BGR2GRAY)
                sz = im1_gray.shape
                warp_mode = cv2.MOTION_EUCLIDEAN 
                warp_matrix = np.eye(2, 3, dtype=np.float32)
                number_of_iterations = 5000;
                termination_eps = 1e-10;
                criteria = (cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, number_of_iterations,  termination_eps)
                (cc, warp_matrix) = cv2.findTransformECC(im1_gray,im2_gray,warp_matrix, warp_mode, criteria)
                im2_aligned = cv2.warpAffine(Slide2, warp_matrix, (sz[1],sz[0]), flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
                stack[row-first_line,0:sz[0],0:sz[1],:] = im2_aligned
                Sli = re.search(Side+'(.*).svs',readexcel.Filelist[0])
                if 'r' in Sli.group(1):
                    index = Sli.group(1).index('r') 
                    slidenumber = Sli.group(1)[0:index]
                else:
                    slidenumber = Sli.group(1)
                
                Transformed_X_cor_HC,Transformed_Y_cor_HC = calculateaffinetransformation(slidenumber,HC_slide,HC_local_X_Cor,HC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_HC,Transformed_Y_cor_HC)
                Transformed_X_cor_SGN,Transformed_Y_cor_SGN = calculateaffinetransformation(slidenumber,SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_SGN,Transformed_Y_cor_SGN)
                Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
                # Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
        
            for n,row in enumerate(np.arange(MidModiolar_line+1,first_line+1,-1).astype(int)):
                direction = -1
                print(row)
                readexcel(ReadfromDatasource,row)
                if n == 0:
                    x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                    Slide1 = read_slide.image
                else:    
                    Slide1 = im2_aligned
                readexcel(ReadfromDatasource,row-1)
                x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                Slide2 = read_slide.image
                im1_gray = cv2.cvtColor(Slide1,cv2.COLOR_BGR2GRAY)
                im2_gray = cv2.cvtColor(Slide2,cv2.COLOR_BGR2GRAY)
                sz = im1_gray.shape
                warp_mode = cv2.MOTION_EUCLIDEAN 
                warp_matrix = np.eye(2, 3, dtype=np.float32)
                number_of_iterations = 5000;
                termination_eps = 1e-10;
                criteria = (cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, number_of_iterations,  termination_eps)
                (cc, warp_matrix) = cv2.findTransformECC (im1_gray,im2_gray,warp_matrix, warp_mode, criteria)
                im2_aligned = cv2.warpAffine(Slide2, warp_matrix, (sz[1],sz[0]), flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
                
                stack[row-first_line,0:sz[0],0:sz[1],:] = im2_aligned
                Sli = re.search(Side+'(.*).svs',readexcel.Filelist[0])
                if 'r' in Sli.group(1):
                    index = Sli.group(1).index('r') #number+'r' is rescanned image
                    slidenumber = Sli.group(1)[0:index]
                else:
                    slidenumber = Sli.group(1)
                Transformed_X_cor_HC,Transformed_Y_cor_HC = calculateaffinetransformation(slidenumber,HC_slide,HC_local_X_Cor,HC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_HC,Transformed_Y_cor_HC)
                Transformed_X_cor_SGN,Transformed_Y_cor_SGN = calculateaffinetransformation(slidenumber,SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_SGN,Transformed_Y_cor_SGN)
                Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
            To_list =np.array( [Transformed_X_cor_HC,Transformed_Y_cor_HC,HC_slide,HC_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for HC v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for HC v6.xlsx'))  
            
            To_list =np.array( [Transformed_X_cor_RC,Transformed_Y_cor_RC,RC_slide,RC_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for RC v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for RC v6.xlsx'))  
            
            To_list =np.array( [Transformed_X_cor_SGN,Transformed_Y_cor_SGN,SGN_slide,SGN_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for SGN v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for SGN v6.xlsx'))
            successin20x = successin20x + [Row_index]
    except:
        cannotconverge = cannotconverge+[Row_index]
        print(' error')
        print(cannotconverge)
        pass
 
#%% 
Onlyfailthefirstfewslides = []
cannotconvergeon20x = []
Row_INDEX = [94,48]
for Row_index in Row_INDEX:
    # try:
        Indextoaddress  = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Code and data  for figures v10\Index to addressbook.xlsx')
        ReadIndextoaddress = Indextoaddress.active
        traynumber = ReadIndextoaddress.cell(Row_index,column = 1).value
        Side = ReadIndextoaddress.cell(Row_index,column = 2).value
        print('working on ',str(traynumber))
        firstslide = ReadIndextoaddress.cell(Row_index,column = 6).value
        first_line = ReadIndextoaddress.cell(Row_index,column = 3).value # the first line in excel to read from 
        MidModiolar_line = ReadIndextoaddress.cell(Row_index,column = 4).value
        last_line = ReadIndextoaddress.cell(Row_index,column = 5).value # the last line in excel to read
        scale = 1/1994
        zoomout = 20
        Datasource = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Code and data  for figures v10\Address book for cases in my database.xlsx')
        ReadfromDatasource = Datasource.active
        rootdatapathformask = r'\\apollo\research\ENT\Liberman\Peizhe Wu\01 Human study WPZ\01 Human Ear Raw Data\Human temporal bone 3D stack'
        datapathformask = os.path.join(rootdatapathformask,'Tray'+str(traynumber),'Mask')
        datapathforcoordinate = os.path.join(rootdatapathformask,'Tray'+str(traynumber))
        localpath = os.path.join('D:\SGC project data\Human temporal bone 3D stack','Tray'+str(traynumber))
        newdatapathformask= os.path.join(rootdatapathformask,'Tray'+str(traynumber),'Transformed Mask')
        Allmaskfiles = find('*merge_mask.tif', datapathformask)
        if os.path.isfile(os.path.join(datapathforcoordinate,'xyz coordinates for HC v4.xlsx')):
    
            HCdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for HC.xlsx'))
            HCdatafromdatasource = HCdatasource.active
            HC_slide,HC_local_X_Cor,HC_local_Y_Cor,HC_local_Z_Cor = [Readcoordinate(HCdatafromdatasource,'D',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'B',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'C',[],1),
                                                      Readcoordinate(HCdatafromdatasource,'G',[],1)]
            Transformed_X_cor_HC,Transformed_Y_cor_HC = HC_local_X_Cor.copy(),HC_local_Y_Cor.copy()
          
                        
                        
            RCdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for RC.xlsx'))
            RCdatafromdatasource = RCdatasource.active
            RC_slide,RC_local_X_Cor,RC_local_Y_Cor,RC_local_Z_Cor = [Readcoordinate(RCdatafromdatasource,'D',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'B',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'C',[],1),
                                                      Readcoordinate(RCdatafromdatasource,'G',[],1)]
            Transformed_X_cor_RC,Transformed_Y_cor_RC = RC_local_X_Cor.copy(),RC_local_Y_Cor.copy()
            
            SGNdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for SGN.xlsx'))
            SGNdatafromdatasource = SGNdatasource.active
            SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,SGN_local_Z_Cor = [Readcoordinate(SGNdatafromdatasource,'D',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'B',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'C',[],1),
                                                  Readcoordinate(SGNdatafromdatasource,'G',[],1)]
            Transformed_X_cor_SGN,Transformed_Y_cor_SGN = SGN_local_X_Cor.copy(),SGN_local_Y_Cor.copy()
            
            # Implantdatasource = load_workbook(os.path.join(datapathforcoordinate,'xyz coordinates for implant.xlsx'))
            # Implantdatafromdatasource = Implantdatasource.active
            # Implant_slide,Implant_local_X_Cor,Implant_local_Y_Cor,Implant_local_Z_Cor = [Readcoordinate(SGNdatafromdatasource,'D',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'B',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'C',[],1),
            #                                       Readcoordinate(SGNdatafromdatasource,'G',[],1)]
            # Transformed_X_cor_Implant,Transformed_Y_cor_Implant = Implant_local_X_Cor.copy(),Implant_local_Y_Cor.copy()
        
            for n,row in enumerate(np.arange(MidModiolar_line,last_line,1).astype(int)):    
            # for n,row in enumerate(np.arange(3520,last_line,1).astype(int)):
                direction = 1
                print(row)
                readexcel(ReadfromDatasource,row)
                if n == 0:
                    x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                    Slide1 = read_slide.image
                    stack =np.zeros([50,int(math.ceil(height_T/zoomout)),int(math.ceil(width_T/zoomout)),3],dtype=np.uint8)
                else:    
                    Slide1 = im2_aligned
                readexcel(ReadfromDatasource,row+1)
                x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
        
                read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                Slide2 = read_slide.image
                
                im1_gray = cv2.cvtColor(Slide1,cv2.COLOR_BGR2GRAY)
                im2_gray = cv2.cvtColor(Slide2,cv2.COLOR_BGR2GRAY)
                sz = im1_gray.shape
                warp_mode = cv2.MOTION_EUCLIDEAN 
                warp_matrix = np.eye(2, 3, dtype=np.float32)
                number_of_iterations = 5000;
                termination_eps = 1e-10;
                criteria = (cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, number_of_iterations,  termination_eps)
                (cc, warp_matrix) = cv2.findTransformECC(im1_gray,im2_gray,warp_matrix, warp_mode, criteria)
                im2_aligned = cv2.warpAffine(Slide2, warp_matrix, (sz[1],sz[0]), flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
                stack[row-first_line,0:sz[0],0:sz[1],:] = im2_aligned
                Sli = re.search(Side+'(.*).svs',readexcel.Filelist[0])
                if 'r' in Sli.group(1):
                    index = Sli.group(1).index('r') 
                    slidenumber = Sli.group(1)[0:index]
                else:
                    slidenumber = Sli.group(1)
                
                Transformed_X_cor_HC,Transformed_Y_cor_HC = calculateaffinetransformation(slidenumber,HC_slide,HC_local_X_Cor,HC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_HC,Transformed_Y_cor_HC)
                Transformed_X_cor_SGN,Transformed_Y_cor_SGN = calculateaffinetransformation(slidenumber,SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_SGN,Transformed_Y_cor_SGN)
                Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
                # Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
        
            for n,row in enumerate(np.arange(MidModiolar_line+1,first_line+1,-1).astype(int)):
                direction = -1
                print(row)
                readexcel(ReadfromDatasource,row)
                if n == 0:
                    x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                    Slide1 = read_slide.image
                else:    
                    Slide1 = im2_aligned
                readexcel(ReadfromDatasource,row-1)
                x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
                read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T, width=width_T,height=height_T,scale=zoomout)
                Slide2 = read_slide.image
                im1_gray = cv2.cvtColor(Slide1,cv2.COLOR_BGR2GRAY)
                im2_gray = cv2.cvtColor(Slide2,cv2.COLOR_BGR2GRAY)
                sz = im1_gray.shape
                warp_mode = cv2.MOTION_EUCLIDEAN 
                warp_matrix = np.eye(2, 3, dtype=np.float32)
                number_of_iterations = 5000;
                termination_eps = 1e-10;
                criteria = (cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, number_of_iterations,  termination_eps)
                (cc, warp_matrix) = cv2.findTransformECC (im1_gray,im2_gray,warp_matrix, warp_mode, criteria)
                im2_aligned = cv2.warpAffine(Slide2, warp_matrix, (sz[1],sz[0]), flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
                
                stack[row-first_line,0:sz[0],0:sz[1],:] = im2_aligned
                Sli = re.search(Side+'(.*).svs',readexcel.Filelist[0])
                if 'r' in Sli.group(1):
                    index = Sli.group(1).index('r') #number+'r' is rescanned image
                    slidenumber = Sli.group(1)[0:index]
                else:
                    slidenumber = Sli.group(1)
                Transformed_X_cor_HC,Transformed_Y_cor_HC = calculateaffinetransformation(slidenumber,HC_slide,HC_local_X_Cor,HC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_HC,Transformed_Y_cor_HC)
                Transformed_X_cor_SGN,Transformed_Y_cor_SGN = calculateaffinetransformation(slidenumber,SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_SGN,Transformed_Y_cor_SGN)
                Transformed_X_cor_RC,Transformed_Y_cor_RC = calculateaffinetransformation(slidenumber,RC_slide,RC_local_X_Cor,RC_local_Y_Cor,zoomout,Slide2,warp_matrix,Transformed_X_cor_RC,Transformed_Y_cor_RC)
            To_list =np.array( [Transformed_X_cor_HC,Transformed_Y_cor_HC,HC_slide,HC_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for HC v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for HC v6.xlsx'))  
            
            To_list =np.array( [Transformed_X_cor_RC,Transformed_Y_cor_RC,RC_slide,RC_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for RC v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for RC v6.xlsx'))  
            
            To_list =np.array( [Transformed_X_cor_SGN,Transformed_Y_cor_SGN,SGN_slide,SGN_local_Z_Cor])
            save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
            save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for SGN v6.xlsx'))  
            save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for SGN v6.xlsx'))
            Onlyfailthefirstfewslides = Onlyfailthefirstfewslides + [Row_index]
            print('error can be overwritten')
            imwrite(os.path.join(datapathforcoordinate,'image stack after affine tranformation v2.tif'), stack, imagej=True)  
            imwrite(os.path.join(localpath,'image stack after affine tranformationv2.tif'), stack, imagej=True)  

    # except:
    #     print(' error')
    #     print(cannotconvergeon20x)
    #     if int(slidenumber) < min(HC_slide):
    #         To_list =np.array( [Transformed_X_cor_HC,Transformed_Y_cor_HC,HC_slide,HC_local_Z_Cor])
    #         save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
    #         save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for HC v6.xlsx'))  
    #         save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for HC v6.xlsx'))  
            
    #         To_list =np.array( [Transformed_X_cor_RC,Transformed_Y_cor_RC,RC_slide,RC_local_Z_Cor])
    #         save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
    #         save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for RC v6.xlsx'))  
    #         save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for RC v6.xlsx'))  
            
    #         To_list =np.array( [Transformed_X_cor_SGN,Transformed_Y_cor_SGN,SGN_slide,SGN_local_Z_Cor])
    #         save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','Zcor'])    
    #         save_to_file.to_excel(os.path.join(datapathforcoordinate,'xyz coordinates for SGN v6.xlsx'))  
    #         save_to_file.to_excel(os.path.join(localpath,'xyz coordinates for SGN v6.xlsx'))
    #         Onlyfailthefirstfewslides = Onlyfailthefirstfewslides + [Row_index]
    #         print('error can be overwritten')
    #     else:
    #         cannotconvergeon20x = cannotconvergeon20x+[Row_index]
    #     pass
