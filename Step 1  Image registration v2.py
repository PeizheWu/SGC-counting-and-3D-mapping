# -*- coding: utf-8 -*-
"""
Created on Tue Oct 12 13:27:18 2021

@author: WUPEIZH

clean up image registratio code 
"""
#%% Import packages
import slideio
#import matplotlib.pyplot as plt
import cv2
import os
import math
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from tifffile import imwrite
import matplotlib.pyplot as plt
import re

# %% get all the file name
DIRECTORY = r'\\apollo\research\ENT\OtopathWholeScan\AT2-Images\MEEI-HTB20x\H1437R-Cisplatin-Rescanned'
Filelist = os.listdir(DIRECTORY)
Filelist = pd.DataFrame(Filelist)
Filelist.to_clipboard(sep=',',index=False)
#%% enter parameters
Row_index = 84
scale = 1/1994
zoomout = 40.0
Indextoaddress  = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Index to addressbook.xlsx')
ReadIndextoaddress = Indextoaddress.active
traynumber = ReadIndextoaddress.cell(Row_index,column = 1).value
savetopath = r'\\apollo\research\ENT\Liberman\Peizhe Wu\01 Human study WPZ\01 Human Ear Raw Data\Human temporal bone 3D stack\Tray'+str(traynumber)
if not os.path.exists(savetopath):
    os.makedirs(savetopath)
Side = ReadIndextoaddress.cell(Row_index,column = 2).value
firstslide = ReadIndextoaddress.cell(Row_index,column = 6).value
first_line = ReadIndextoaddress.cell(Row_index,column = 3).value 
MidModiolar_line = ReadIndextoaddress.cell(Row_index,column = 4).value
last_line = ReadIndextoaddress.cell(Row_index,column = 5).value
Addressbook = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Address book for cases in my database.xlsx')

#%% define functions for image processing
def read_slide(full,xstart = 40443,ystart =16254,width = 4995,height = 3333,scale=1):
    slide = slideio.open_slide(path=full,driver = 'SVS')
    scene = slide.get_scene(0)
    read_slide.image = scene.read_block(rect=(xstart,ystart,width,height),size = (math.ceil(width/scale),math.ceil(height/scale)),slices=(0,0))
    print("Image imported from "+full)
    
def readexcel(ReadfromDatasource,rows):
    readexcel.Filelist = [ReadfromDatasource.cell(row = rows,column = 2).value]
    readexcel.X_cor = [ReadfromDatasource.cell(row = rows,column = 3).value]
    readexcel.Y_cor = [ReadfromDatasource.cell(row = rows,column = 4).value]
    readexcel.Width_cor =  [ReadfromDatasource.cell(row = rows,column = 5).value]
    readexcel.Height_cor =  [ReadfromDatasource.cell(row = rows,column = 6).value]
    readexcel.folder = [ReadfromDatasource.cell(row = rows,column = 8).value]
    readexcel.Tray = [ReadfromDatasource.cell(row = rows,column = 1).value]

#%% Define functions for 3D recon
def savecor(listofcor,remove=[],scale = 1):
    Arrayoflist = np.asarray(listofcor)
    Arrayoflist= Arrayoflist*scale
    Arrayoflist = np.delete(Arrayoflist, remove)
    tofile = np.array(Arrayoflist).T.tolist()
    return tofile

def rescale(listofcor,scale = 1,Firstslide=0):
    Arrayoflist = np.asarray(listofcor)
    Arrayoflist= (Arrayoflist-Firstslide)*scale
    Arrayoflist= Arrayoflist.astype(np.float)
    Arrayoflist = Arrayoflist.tolist()
    return Arrayoflist

def Readcoordinate(datasource,column,coordinate,readfromrow):
    thecolumn = datasource[column]
    for cell in thecolumn[readfromrow:]:
        coordinate = coordinate+[cell.value]
    return coordinate    

#%% Define functions to calculate  distance from base
def calculatedistance(coordinate):
    size = len(coordinate[0])
    distancefrombase = [0]
    accumulatedDistance=0
    for loc in range(0,size-1):
        x1,y1,z1 = coordinate[0][loc],coordinate[1][loc],coordinate[2][loc]
        x2,y2,z2 =  coordinate[0][loc+1],coordinate[1][loc+1],coordinate[2][loc+1]
        dist = np.linalg.norm(np.array((x1,y1,z1))-np.array((x2,y2,z2)))
        accumulatedDistance = accumulatedDistance +dist
        distancefrombase = distancefrombase +[accumulatedDistance]    
    percentDfromBase = distancefrombase/accumulatedDistance*100
    return percentDfromBase,distancefrombase
    
    
def matchinnervation(HCCoordinate,SGNCoordinate,PercenD_HC,PercenD_SGN):
    size_HC,size_SGN =[len(HCCoordinate[0]),len(SGNCoordinate[0])]
    matchHC=[]
    indexformatchedSGN=[]
    indexformatchedHC=[]
    matchSGN = []    
    for loc_HC in range(0,size_HC):
        distancetoSGN=[]
        xHC,yHC,zHC= HCCoordinate[0][loc_HC],HCCoordinate[1][loc_HC],HCCoordinate[2][loc_HC]
        for loc_SGN in range(0,size_SGN-1):
            xSGN,ySGN,zSGN= SGNCoordinate[0][loc_SGN],SGNCoordinate[1][loc_SGN],SGNCoordinate[2][loc_SGN]
            dist = np.linalg.norm(np.array((xHC,yHC,zHC))-np.array((xSGN,ySGN,zSGN)))
            distancetoSGN = distancetoSGN +[dist]        
        index_min = np.argmin(np.asarray(distancetoSGN))
        indexformatchedSGN = indexformatchedSGN+[index_min]
        matchSGN =matchSGN+[PercenD_SGN[index_min]]
    for loc_SGN in range(0,size_SGN):
        distancetoHC=[]
        xSGN,ySGN,zSGN= SGNCoordinate[0][loc_SGN],SGNCoordinate[1][loc_SGN],SGNCoordinate[2][loc_SGN]
        for loc_HC in range(0,size_HC-1):
            xHC,yHC,zHC= HCCoordinate[0][loc_HC],HCCoordinate[1][loc_HC],HCCoordinate[2][loc_HC]
            dist = np.linalg.norm(np.array((xHC,yHC,zHC))-np.array((xSGN,ySGN,zSGN)))
            distancetoHC = distancetoHC +[dist]        
        index_min = np.argmin(np.asarray(distancetoHC))
        matchHC = matchHC+ [PercenD_HC[index_min]]
        indexformatchedHC = indexformatchedHC+[index_min]
    
    return matchHC,matchSGN,indexformatchedSGN,indexformatchedHC


def determinefrequencyforSGN(RCCoordinate,SGNCoordinate):
    Percent=np.asarray(RCCoordinate[3])
    RC = np.array(RCCoordinate).T.tolist()
    SGN = np.array(SGNCoordinate).T.tolist()
    matchedPercentageRC=[]
    for ind,SGN_COR in enumerate(SGN):
        x1,y1,z1 = [SGN_COR[0],SGN_COR[1],SGN_COR[2]]
        Alldist,index=100,100
        for ind,RC_COR in enumerate(RC):
            x2,y2,z2 = [RC_COR[0],RC_COR[1],RC_COR[2]]
            dist = np.linalg.norm(np.array((x1,y1,z1))-np.array((x2,y2,z2)))
            index = ind if Alldist > dist else index
            Alldist = dist if Alldist > dist else Alldist
        matchedPercentageRC = matchedPercentageRC+ [Percent[index]]
    return matchedPercentageRC
        # print(matchedPercentageRC)
                                         
        
#%% define function for image registration 

def imageregistration(ReadfromAddressbook,row=1,Direction = 'down',Fullwidth=1,Fullheight=1):
    if Direction =='down':
        direction = 1 
    else:
        direction = -1 
    readexcel(ReadfromAddressbook,row)
    width_checksize,height_checksize = slideio.open_slide(path = os.path.join(readexcel.folder[0],readexcel.Filelist[0]),driver = 'SVS').get_scene(0).size
    if width_checksize >0.70*Fullwidth and height_checksize >0.70*Fullheight:
        x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]         
        print(x_cor_T,y_cor_T,width_T,height_T)
    else:
        readexcel(ReadfromAddressbook,row+2*(-direction))
        width_checksize,height_checksize = slideio.open_slide(path = os.path.join(readexcel.folder[0],readexcel.Filelist[0]),driver = 'SVS').get_scene(0).size
        if width_checksize >0.70*Fullwidth and height_checksize >0.70*Fullheight:
            x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
        else:
            x_cor_T,y_cor_T,width_T,height_T = [0,0,width_checksize,height_checksize]
            print('Two neighbouring small image')
            print(width_checksize, height_checksize)

    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T,ystart =y_cor_T,width=width_T,height=height_T,scale=zoomout)
    template = read_slide.image

    readexcel(ReadfromAddressbook,row+direction)
    width_checksize,height_checksize = slideio.open_slide(path = os.path.join(readexcel.folder[0],readexcel.Filelist[0]),driver = 'SVS').get_scene(0).size
    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = 0,ystart =0, width=width_checksize,height=height_checksize,scale=zoomout)
    Source = read_slide.image

    # plt.imshow(Source)
    # plt.imshow(template)

    method = eval('cv2.TM_CCOEFF_NORMED')
    background = Source.copy()
    crosscorrelation = cv2.matchTemplate(background,template,method)
    min_val, max_val, min_loc, top_left = cv2.minMaxLoc(crosscorrelation)
    
    # background[top_left[1]:top_left[1]+math.ceil(height_T/zoomout),top_left[0]:top_left[0]+math.ceil(width_T/zoomout),0:3]=template
    # plt.imshow(background)
    y_cor_N,x_cor_N = (top_left[1]*zoomout,top_left[0]*zoomout)
    return y_cor_N,x_cor_N,width_T,height_T

#%% define function for get x,y,z coordinate 
def ProcessMouseEvent(event,x,y,flags, params):
    global z_loc,ix,iy,img,stacksize,x_list, y_list,z_list
    # mousewheel event code is 10
    if event == 10:
        #sign of the flag shows direction of mousewheel  
        if flags > 0:
            z_loc+=1
            print(z_loc)
            if z_loc-(stacksize-1)>0:
                z_loc+=-1
            img = Stack[z_loc,:,:,:]
            cv2.imshow('image',img)
        else:
            z_loc+=-1
            if z_loc<0:
                z_loc+=1
            img = Stack[z_loc,:,:,:]
            cv2.imshow('image',img)
            print(z_loc)
    if event == cv2.EVENT_LBUTTONDOWN :
        cv2.circle(img,(x,y),5,(255,0,0),-1)
        x_list,y_list,z_list = [x_list + [x],y_list + [y],z_list + [slidenumber[z_loc]]]

#%% full 3D recon 
ReadfromAddressbook = Addressbook.active
slidenumber = []
for row in np.linspace(first_line,last_line,(last_line-first_line+1)).astype(int):
    Slidename = ReadfromAddressbook.cell(row,column = 2).value
    Sli = re.search(Side+'(.*).svs',Slidename)
    if 'r' in Sli.group(1):
        index = Sli.group(1).index('r') #number+'r' is rescanned image
        slidenumber = slidenumber + [Sli.group(1)[0:index]]
    else:
        slidenumber = slidenumber + [Sli.group(1)]
            

readexcel(ReadfromAddressbook,MidModiolar_line)   
Fullwidth,Fullheight= slideio.open_slide(path = os.path.join(readexcel.folder[0],readexcel.Filelist[0]),driver = 'SVS').get_scene(0).size

for row in np.linspace(MidModiolar_line,last_line-1,(last_line-MidModiolar_line)).astype(int):
    y_cor_N,x_cor_N,width_T,height_T = imageregistration(ReadfromAddressbook,row=row,Direction = 'down',Fullwidth=Fullwidth,Fullheight=Fullheight)
    ReadfromAddressbook.cell(row = row+1,column = 3).value = x_cor_N
    ReadfromAddressbook.cell(row = row+1,column = 4).value = y_cor_N
    ReadfromAddressbook.cell(row = row+1,column = 5).value = width_T
    ReadfromAddressbook.cell(row = row+1,column = 6).value = height_T
   
for row in np.linspace(MidModiolar_line,first_line+1,MidModiolar_line-first_line).astype(int):
    y_cor_N,x_cor_N,width_T,height_T = imageregistration(ReadfromAddressbook,row=row,Direction = 'up',Fullwidth=Fullwidth,Fullheight=Fullheight)  
    ReadfromAddressbook.cell(row = row-1,column = 3).value = x_cor_N
    ReadfromAddressbook.cell(row = row-1,column = 4).value = y_cor_N
    ReadfromAddressbook.cell(row = row-1,column = 5).value = width_T
    ReadfromAddressbook.cell(row = row-1,column = 6).value = height_T

Addressbook.save(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Address book for cases in my database.xlsx')
#%% create 3D stack 
ReadfromAddressbook = Addressbook.active
slidenumber = []
for row in np.linspace(first_line,last_line,(last_line-first_line+1)).astype(int):
    Slidename = ReadfromAddressbook.cell(row,column = 2).value
    Sli = re.search(Side+'(.*).svs',Slidename)
    if 'r' in Sli.group(1):
        index = Sli.group(1).index('r') #number+'r' is rescanned image
        slidenumber = slidenumber + [Sli.group(1)[0:index]]
    else:
        slidenumber = slidenumber + [Sli.group(1)]

readexcel(ReadfromAddressbook,first_line)
width_T,height_T = [int(readexcel.Width_cor[0]), int(readexcel.Height_cor[0])]
Stack =np.zeros((last_line-first_line+1,math.ceil(height_T/zoomout),math.ceil(width_T/zoomout),3)).astype(np.uint8)
for n,row in enumerate (np.linspace(first_line,last_line,(last_line-first_line+1)).astype(int)):
    readexcel(ReadfromAddressbook,row)
    x_cor_T,y_cor_T,width_T,height_T = [int(readexcel.X_cor[0]),int(readexcel.Y_cor[0]), int(readexcel.Width_cor[0]),int(readexcel.Height_cor[0])]
    read_slide(os.path.join(readexcel.folder[0],readexcel.Filelist[0]),xstart = x_cor_T, ystart =y_cor_T,width=width_T,height=height_T,scale=zoomout)
    Stack[n,:,:,:]= read_slide.image.astype(np.uint8)
    
imwrite(os.path.join(savetopath,'stack.tif'), Stack, imagej=True)    
#%% get xyz coordinate for HC 
x_list, y_list,z_list= [[],[],[]]
stacksize = Stack.shape[0]
cv2.namedWindow('image')
cv2.setMouseCallback('image',ProcessMouseEvent)
z_loc=1
while(1):
    k = cv2.waitKey(20) & 0xFF
    if k == 27:
        break
    elif k == ord('a'):
        print(x_list, y_list,z_list)
cv2.destroyAllWindows()  

x_list_HC,y_list_HC,z_list_HC = [x_list,y_list,np.asarray(z_list).astype(np.integer).tolist()]
#%% get xyz coordinate for RC and save the stack 
x_list, y_list,z_list= [[],[],[]]
stacksize = Stack.shape[0]
cv2.namedWindow('image')
cv2.setMouseCallback('image',ProcessMouseEvent)
z_loc=1
while(1):
    k = cv2.waitKey(20) & 0xFF
    if k == 27:
        break
    elif k == ord('a'):
        print(x_list, y_list,z_list)
cv2.destroyAllWindows()  

x_list_RC,y_list_RC,z_list_RC = [x_list,y_list,np.asarray(z_list).astype(np.integer).tolist()]

# imwrite(os.path.join(savetopath,'stack.tif'), Stack, imagej=True)    
     

#%% clean up the data and save to excel 
remove = []
x_HC,y_HC,z_HC= [savecor(x_list_HC,[],zoomout),savecor(y_list_HC,[],zoomout),savecor(z_list_HC,[])]  
x_RC,y_RC,z_RC= [savecor(x_list_RC,[],zoomout),savecor(y_list_RC,[],zoomout),savecor(z_list_RC,[])]
x_HC_toplot,y_HC_toplot,z_HC_toplot= [rescale(x_HC,scale),rescale(y_HC,scale),rescale(z_HC,2/100,firstslide)]  
x_RC_toplot,y_RC_toplot,z_RC_toplot= [rescale(x_RC,scale),rescale(y_RC,scale),rescale(z_RC,2/100,firstslide)]   

HCCoordinate = [x_HC_toplot,y_HC_toplot,z_HC_toplot]
RCCoordinate = [x_RC_toplot,y_RC_toplot,z_RC_toplot]
PercenD_HC,AbsD_HC = calculatedistance(HCCoordinate)
PercenD_RC,AbsD_RC = calculatedistance(RCCoordinate)
matchHC,matchRC,index_RC,index_HC= matchinnervation(HCCoordinate,RCCoordinate,PercenD_HC,PercenD_RC)
    
To_list =np.array( [x_RC,y_RC,z_RC,x_RC_toplot,y_RC_toplot,z_RC_toplot,PercenD_RC,AbsD_RC,matchHC,index_HC])
save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','x_SGN', 'y_SGN','z_SGN','PercetfromBase',
                                                  'distance from base mm','matchedHCpercentagefrombase','indexmatchedHC'])
save_to_file.to_excel(os.path.join(savetopath,'xyz coordinates for RC.xlsx'))  


To_list =np.array( [x_HC,y_HC,z_HC,x_HC_toplot,y_HC_toplot,z_HC_toplot,PercenD_HC,AbsD_HC,matchRC,index_RC])
save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','x_HC', 'y_HC','z_HC','PercetfromBase',
                                                  'distance from base mm','matchedSGNPercentfrombase','indexmatchedRC'])
save_to_file.to_excel(os.path.join(savetopath,'xyz coordinates for HC.xlsx'))                                       

imwrite(os.path.join(savetopath,'stack.tif'), Stack, imagej=True)    
#%% plot frequency map and save it

fig, ax = plt.subplots(constrained_layout=True)
ax.plot(PercenD_HC,matchRC)
ax.set_xlabel(' HC percent from base')
ax.set_ylabel('matched SGN per from base')
ax.set_title('Frequency map of Rosenthal cannal')

plt.savefig(os.path.join(savetopath,'Tray'+str(traynumber)+'.map.png'))       


#%% get xyz coordinate for Implant and save the data 
x_list, y_list,z_list= [[],[],[]]
stacksize = Stack.shape[0]
cv2.namedWindow('image')
cv2.setMouseCallback('image',ProcessMouseEvent)
z_loc=1
while(1):
    k = cv2.waitKey(20) & 0xFF
    if k == 27:
        break
    elif k == ord('a'):
        print(x_list, y_list,z_list)
cv2.destroyAllWindows()  

x_list_implant,y_list_implant,z_list_implant = [x_list,y_list,np.asarray(z_list).astype(np.integer).tolist()]

# imwrite(os.path.join(savetopath,'stack.tif'), Stack, imagej=True)    

remove = []

x_HC,y_HC,z_HC= [savecor(x_list_implant,[],zoomout),savecor(y_list_implant,[],zoomout),savecor(z_list_implant,[])]  
x_HC_toplot,y_HC_toplot,z_HC_toplot= [rescale(x_HC,scale),rescale(y_HC,scale),rescale(z_HC,2/100,firstslide)]  
HCCoordinate = [x_HC_toplot,y_HC_toplot,z_HC_toplot]
    
To_list =np.array( [x_HC,y_HC,z_HC,x_HC_toplot,y_HC_toplot,z_HC_toplot])
save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','x_HC', 'y_HC','z_HC'])
save_to_file.to_excel(os.path.join(savetopath,'xyz coordinates for implant.xlsx'))    
