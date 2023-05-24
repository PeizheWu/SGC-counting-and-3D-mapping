# -*- coding: utf-8 -*-
"""
Created on Sat Aug 28 17:38:06 2021

@author: WUPEIZH
Do multiple ML model at a time 
"""

#%% import package
import slideio
#import matplotlib.pyplot as plt

import cv2
import os
import math
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import time
import torch
import matplotlib.pyplot as plt
torch.cuda.empty_cache()
from torch.utils.data import DataLoader
from torch.utils.data import Dataset as BaseDataset
import segmentation_models_pytorch as smp # so this is a library
import albumentations as albu
# Albumentations is a Python library for image augmentation. 
# Image augmentation is used in deep learning and computer vision tasks to increase the quality of trained models. 
# The purpose of image augmentation is to create new training samples from the existing data.
#%% define helper functions 
def read_slide(full,xstart = 40443,ystart =16254,width = 4995,height = 3333):
    slide = slideio.open_slide(path=full,driver = 'SVS')
    scene = slide.get_scene(0)
    read_slide.image = scene.read_block(rect=(xstart,ystart,width,height),size = (0,0),slices=(0,0))
    print("Image imported from "+full)

def crop_image(image,cropwidth=2400,cropheight=3600,cropxstart = 0,cropystart = 0,name = '878.SVS',roundn = 0,path = r'C:\Users\WUPEIZH\Desktop'):
    cropped = image[cropystart:cropystart+cropheight,cropxstart:cropxstart+cropwidth]
    # plt.imshow(cropped)    
    background = np.zeros([3600,2400,3]);
    background = background.astype(np.uint8)
    background[0:cropped.shape[0],0:cropped.shape[1]]=cropped
    background_after = np.rot90(background)
    background_after = cv2.cvtColor(background_after,cv2.COLOR_BGR2RGB)
    name = 'segment'+ str(roundn)+'.tif'
    cv2.imwrite(os.path.join(path,name),background_after)
    print("Image written for " + path+name)
    
def find_crop_coordinate(image,cropwidth = 2400,cropheight=3600):
    cropwidth = int(cropwidth)
    cropheight = int(cropheight)
    [imageheight,imagewidth,c]=image.shape
    find_crop_coordinate.NX = math.ceil(imagewidth/cropwidth)
    find_crop_coordinate.NY = math.ceil(imageheight/cropheight)
    #print("number of x tile",find_crop_coordinate.NX, '\n', "number of y tile",find_crop_coordinate.NY)    
#%% define AI helper functions
def visualize(**images):
    """PLot images in one row."""
    n = len(images)
    plt.figure(figsize=(16, 5))
    for i, (name, image) in enumerate(images.items()):
        plt.subplot(1, n, i + 1)
        plt.xticks([])
        plt.yticks([])
        plt.title(' '.join(name.split('_')).title())
        plt.imshow(image)
    plt.show()

# define Datset class
class Dataset(BaseDataset):
    """CamVid Dataset. Read images, apply augmentation and preprocessing transformations.
    
    Args:
        images_dir (str): path to images folder
        masks_dir (str): path to segmentation masks folder
        class_values (list): values of classes to extract from segmentation mask
        augmentation (albumentations.Compose): data transfromation pipeline 
            (e.g. flip, scale, etc.)
        preprocessing (albumentations.Compose): data preprocessing 
            (e.g. noralization, shape manipulation, etc.)
    
    """
    
    CLASSES = ['backgroud','cell']
 
    
    def __init__(
            self, 
            images_dir, 
            masks_dir, 
            classes=None, 
            augmentation=None, 
            preprocessing=None,):
        self.ids = os.listdir(images_dir)
        self.images_fps = [os.path.join(images_dir, image_id) for image_id in self.ids]
        self.masks_fps = [os.path.join(masks_dir, image_id) for image_id in self.ids]
        
        # convert str names to class values on masks
        self.class_values = [self.CLASSES.index(cls.lower()) for cls in classes]
        self.augmentation = augmentation
        self.preprocessing = preprocessing
    
    def __getitem__(self, i):
        
        # read data
        image = cv2.imread(self.images_fps[i])
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        mask = cv2.imread(self.masks_fps[i], 0)
        
        # extract certain classes from mask (e.g. cars)
        masks = [(mask == v) for v in self.class_values]
        mask = np.stack(masks, axis=-1).astype('float')
        
        # apply augmentations
        if self.augmentation:
            sample = self.augmentation(image=image, mask=mask)
            image, mask = sample['image'], sample['mask']
        
        # apply preprocessing
        if self.preprocessing:
            sample = self.preprocessing(image=image, mask=mask)
            image, mask = sample['image'], sample['mask']
            
        return image, mask
        
    def __len__(self):
        return len(self.ids)

# define get_training_augmentation function
def get_training_augmentation():
    train_transform = [

        albu.HorizontalFlip(p=0.5),# 50% that the image will flip horizontally 

        albu.ShiftScaleRotate(scale_limit=0.5, rotate_limit=0, shift_limit=0.1, p=1, border_mode=0),

        albu.PadIfNeeded(min_height=2400, min_width=2400, always_apply=True, border_mode=0), 
        albu.RandomCrop(height=2400, width=2400, always_apply=True), # crop the image into 320 by 320 

        albu.IAAAdditiveGaussianNoise(p=0.2),# add gaussian noise
        albu.IAAPerspective(p=0.5), 

        albu.OneOf(
            [
                albu.CLAHE(p=1),#Apply Contrast Limited Adaptive Histogram Equalization to the input image.
                albu.RandomBrightness(p=1),
                albu.RandomGamma(p=1),
            ],
            p=0.9,
        ),

        albu.OneOf(
            [
                albu.IAASharpen(p=1),
                albu.Blur(blur_limit=3, p=1),
                albu.MotionBlur(blur_limit=3, p=1),
            ],
            p=0.9,
        ),

        albu.OneOf(
            [
                albu.RandomContrast(p=1),
                albu.HueSaturationValue(p=1),
            ],
            p=0.9,
        ),
    ]
    return albu.Compose(train_transform)

# define get_validation_augmentation function
def get_validation_augmentation():
    """Add paddings to make image shape divisible by 32"""
    test_transform = [
        albu.PadIfNeeded(2400,3616) # was (768,960) , try (2880,3600)
    ]
    return albu.Compose(test_transform)

# define to_tensor function 
def to_tensor(x, **kwargs):
    return x.transpose(2, 0, 1).astype('float32')

# define get_preprocessing function
def get_preprocessing(preprocessing_fn):
    """Construct preprocessing transform
    
    Args:
        preprocessing_fn (callbale): data normalization function 
            (can be specific for each pretrained neural network)
    Return:
        transform: albumentations.Compose
    
    """
    
    _transform = [
        albu.Lambda(image=preprocessing_fn),
        albu.Lambda(image=to_tensor, mask=to_tensor),
    ]
    return albu.Compose(_transform)

#%% Choose backbone 
ENCODER = 'resnet34' # Should use Unet, resnet34 encoder and imagenet pretrain weight 
ENCODER_WEIGHTS = 'imagenet'
CLASSES = ['cell']
ACTIVATION = 'sigmoid' # could be None for logits or 'softmax2d' for multicalss segmentation
DEVICE = 'cpu'

# create segmentation model with pretrained encoder
model = smp.Unet(
    encoder_name=ENCODER, 
    encoder_weights=ENCODER_WEIGHTS, 
    classes=len(CLASSES), 
    activation=ACTIVATION,
)

preprocessing_fn = smp.encoders.get_preprocessing_fn(ENCODER, ENCODER_WEIGHTS)

#%% load best model
best_model_red = torch.load(r'C:\Users\WUPEIZH\Dropbox (Partners HealthCare)\04 Customize wholeslide for AI\Models\05 M07292021Trial1.pth',map_location=torch.device('cpu'))
best_model_blue = torch.load(r'C:\Users\WUPEIZH\Dropbox (Partners HealthCare)\04 Customize wholeslide for AI\Models\01 M12272020Trial3.pth',map_location=torch.device('cpu'))
# Detect_SGN_model_12-27-2020.trial3.pth is the latest model, trained and saved on GPU, work just as well on CPU. 
#%% import package 
import slideio
#import matplotlib.pyplot as plt
import cv2
import os
import math
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from tifffile import imwrite
import matplotlib.pyplot as plt
import re
import matplotlib.colors as mcolors
import matplotlib.cm as cmx



#%% read slide helper function 

def readexcel(ReadfromDatasource,rows):
    readexcel.Filelist = [ReadfromDatasource.cell(row = rows,column = 2).value]
    readexcel.X_cor = [ReadfromDatasource.cell(row = rows,column = 3).value]
    readexcel.Y_cor = [ReadfromDatasource.cell(row = rows,column = 4).value]
    readexcel.Width_cor =  [ReadfromDatasource.cell(row = rows,column = 5).value]
    readexcel.Height_cor =  [ReadfromDatasource.cell(row = rows,column = 6).value]
    readexcel.folder = [ReadfromDatasource.cell(row = rows,column = 8).value]
    readexcel.Tray = [ReadfromDatasource.cell(row = rows,column = 1).value]

def Readcoordinate(datasource,column,coordinate,readfromrow):
    thecolumn = datasource[column]
    for cell in thecolumn[readfromrow:]:
        coordinate = coordinate+[cell.value]
    return coordinate    


def savecor(listofcor,remove=[],scale = 1,):
    Arrayoflist = np.asarray(listofcor)
    Arrayoflist= Arrayoflist*scale
    Arrayoflist = np.delete(Arrayoflist, remove)
    tofile = np.array(Arrayoflist).T.tolist()
    return tofile

def rescale(listofcor,scale = 1,Firstslide=0):
    Arrayoflist = np.asarray(listofcor)
    Arrayoflist= (Arrayoflist-Firstslide)*scale
    Arrayoflist= Arrayoflist.astype(np.float)
    return Arrayoflist


def calculatedistance(coordinate):
    size = len(coordinate[0])
    distancefrombase = [0]
    accumulatedDistance=0
    for loc in range(0,size-1):
        x1,y1,z1 = coordinate[0][loc],coordinate[1][loc],coordinate[2][loc]
        x2,y2,z2 =  coordinate[0][loc+1],coordinate[1][loc+1],coordinate[2][loc+1]
        dist = np.linalg.norm(np.array((x1,y1,z1))-np.array((x2,y2,z2)))
        accumulatedDistance = accumulatedDistance +dist
        distancefrombase = distancefrombase +[accumulatedDistance]    
    percentDfromBase = distancefrombase/accumulatedDistance*100
    return percentDfromBase,distancefrombase
    
    
    
def matchinnervation(HCCoordinate,SGNCoordinate,PercenD_HC,PercenD_SGN):
    size_HC,size_SGN =[len(HCCoordinate[0]),len(SGNCoordinate[0])]
    matchHC=[]
    indexformatchedSGN=[]
    indexformatchedHC=[]
    matchSGN = []    
    for loc_HC in range(0,size_HC):
        distancetoSGN=[]
        xHC,yHC,zHC= HCCoordinate[0][loc_HC],HCCoordinate[1][loc_HC],HCCoordinate[2][loc_HC]
        for loc_SGN in range(0,size_SGN-1):
            xSGN,ySGN,zSGN= SGNCoordinate[0][loc_SGN],SGNCoordinate[1][loc_SGN],SGNCoordinate[2][loc_SGN]
            dist = np.linalg.norm(np.array((xHC,yHC,zHC))-np.array((xSGN,ySGN,zSGN)))
            distancetoSGN = distancetoSGN +[dist]        
        index_min = np.argmin(np.asarray(distancetoSGN))
        indexformatchedSGN = indexformatchedSGN+[index_min]
        matchSGN =matchSGN+[PercenD_SGN[index_min]]
    for loc_SGN in range(0,size_SGN):
        distancetoHC=[]
        xSGN,ySGN,zSGN= SGNCoordinate[0][loc_SGN],SGNCoordinate[1][loc_SGN],SGNCoordinate[2][loc_SGN]
        for loc_HC in range(0,size_HC-1):
            xHC,yHC,zHC= HCCoordinate[0][loc_HC],HCCoordinate[1][loc_HC],HCCoordinate[2][loc_HC]
            dist = np.linalg.norm(np.array((xHC,yHC,zHC))-np.array((xSGN,ySGN,zSGN)))
            distancetoHC = distancetoHC +[dist]        
        index_min = np.argmin(np.asarray(distancetoHC))
        matchHC = matchHC+ [PercenD_HC[index_min]]
        indexformatchedHC = indexformatchedHC+[index_min]
    
    return matchHC,matchSGN,indexformatchedSGN,indexformatchedHC


def determinefrequencyforSGN(RCCoordinate,SGNCoordinate):
    Percent=np.asarray(RCCoordinate[3])
    RC = np.array(RCCoordinate).T.tolist()
    SGN = np.array(SGNCoordinate).T.tolist()
    matchedPercentageRC=[]
    for ind,SGN_COR in enumerate(SGN):
        x1,y1,z1 = [SGN_COR[0],SGN_COR[1],SGN_COR[2]]
        Alldist,index=100,100
        for ind,RC_COR in enumerate(RC):
            x2,y2,z2 = [RC_COR[0],RC_COR[1],RC_COR[2]]
            dist = np.linalg.norm(np.array((x1,y1,z1))-np.array((x2,y2,z2)))
            index = ind if Alldist > dist else index
            Alldist = dist if Alldist > dist else Alldist
        matchedPercentageRC = matchedPercentageRC+ [Percent[index]]
    return matchedPercentageRC
        # print(matchedPercentageRC)
                                         
#%% define cases to do 

SAVETOPATH,SIDE,FIRSTSLIDE,FIRST_LINE,LAST_LINE=[[],[],[],[],[]]
Indextoaddress  = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Code and data  for figures v10\Index to addressbook.xlsx')
ReadIndextoaddress = Indextoaddress.active
Row_index = [58,104]

for n in Row_index:
    traynumber = ReadIndextoaddress.cell( n,column = 1).value
    SAVETOPATH= SAVETOPATH + [r'\\apollo\research\ENT\Liberman\Peizhe Wu\01 Human study WPZ\01 Human Ear Raw Data\Human temporal bone 3D stack\Tray'+str(traynumber)]
    SIDE = SIDE+ [ReadIndextoaddress.cell(n,column = 2).value]
    FIRSTSLIDE =FIRSTSLIDE + [ReadIndextoaddress.cell(n,column = 6).value]
    FIRST_LINE =FIRST_LINE + [ReadIndextoaddress.cell(n,column = 3).value]
    LAST_LINE = LAST_LINE + [ReadIndextoaddress.cell(n,column = 5).value]
Datasource = load_workbook(r'\\apollo\research\ENT\Liberman\Peizhe Wu\12 Publications\Work in progress\8 SGC technique paper\Code and data  for figures v10\Address book for cases in my database.xlsx')
scale = 1/1994
zoomout = 20.0    

# SAVETOPATH = [r'C:\Users\WUPEIZH\Desktop\Tray1801',r'C:\Users\WUPEIZH\Desktop\Tray1812',r'C:\Users\WUPEIZH\Desktop\Tray1817']
# SIDE = ['R','R','L']
# FIRSTSLIDE = [11,21,91]
# FIRST_LINE = [847,765,730] # the first line in excel to read from 
# LAST_LINE = [886,805,763] # the last line in excel to read

#%% ML
Allcases = [SAVETOPATH,SIDE,FIRSTSLIDE,FIRST_LINE,LAST_LINE]
Allcases = np.array([Allcases]).T.tolist()  
for savetopath,Side,firstslide,first_line,last_line in Allcases:
    savetopath,Side=savetopath[0],Side[0]
    firstslide,first_line,last_line = int(firstslide[0]),int(first_line[0]),int(last_line[0])
    
    #  Datasource = load_workbook(r'C:\Users\WUPEIZH\Dropbox (Partners HealthCare)\04 Customize wholeslide for AI\test model on different trays\Book1.xlsx')
    slidenumber,x_COR_SGNCROP,y_COR_SGNCROP,folder_list,Filelist,Tray_list = [[],[],[],[],[],[]]
    ReadfromDatasource = Datasource.active
    for row in range (first_line,last_line):
    # for row in [first_line]:
        Slidename = ReadfromDatasource.cell(row,column = 2).value
        Sli = re.search(Side+'(.*).svs',Slidename)
        if 'r' in Sli.group(1):
            index = Sli.group(1).index('r') #number+'r' is rescanned image
            slidenumber = slidenumber + [Sli.group(1)[0:index]]
        else:
            slidenumber = slidenumber + [Sli.group(1)]
        
        
        x_COR_SGNCROP =x_COR_SGNCROP+ [ReadfromDatasource.cell(row,column = 3).value]
        y_COR_SGNCROP =y_COR_SGNCROP+ [ReadfromDatasource.cell(row,column = 4).value]
        folder_list = folder_list+[ReadfromDatasource.cell(row,column = 8).value]
        Filelist = Filelist + [ReadfromDatasource.cell(row ,column = 2).value]
        Tray_list = Tray_list + [ReadfromDatasource.cell(row ,column = 1).value]   
    
    
    RCdatasource = load_workbook(os.path.join(savetopath,'xyz coordinates for RC.xlsx'))
    RCdatafromdatasource = RCdatasource.active
    SGN_slide,SGN_local_X_Cor,SGN_local_Y_Cor = [Readcoordinate(RCdatafromdatasource,'D',[],1),
                                                 Readcoordinate(RCdatafromdatasource,'B',[],1),
                                                 Readcoordinate(RCdatafromdatasource,'C',[],1)]
    
    SGN_slide = np. array(SGN_slide)
    uniqueSGNSlide = np.unique(SGN_slide)
    x_min_cor,y_min_cor,x_max_cor,y_max_cor,x_canvas,y_canvas=[[],[],[],[],[],[]]
    Tray_crop_list,Folder_crop_list,File_crop_list=[[],[],[]]
    
    a,b,c=[0,0,0]
    for slides in uniqueSGNSlide:   
        c+=1
        x_cor,y_cor=[[],[]]
        print('c',c)
        
        for n,allslide in enumerate(SGN_slide):
            if allslide == slides:
                x_cor, y_cor= [x_cor+ [SGN_local_X_Cor[n]], y_cor+ [SGN_local_Y_Cor[n]]]
        a+=1
        print('a',a)
        x_min_cor,y_min_cor,x_max_cor,y_max_cor= [x_min_cor+[min(x_cor)],y_min_cor+[min(y_cor)],x_max_cor+[max(x_cor)],y_max_cor+[max(y_cor)]]
            
        for m,slidefromtotalstack in enumerate(slidenumber):
            if int(slidefromtotalstack)==slides:
                b+=1
                print('b',b)
                print(slidefromtotalstack)
                x_canvas,y_canvas = [x_canvas + [x_COR_SGNCROP[m]],y_canvas + [y_COR_SGNCROP[m]]]
                Tray_crop_list = Tray_crop_list+[Tray_list[m]]
                Folder_crop_list = Folder_crop_list +[folder_list[m]]
                File_crop_list = File_crop_list+[Filelist[m]]
                # print(m)
    
            
    # X_adjusted_Cor_min,X_adjusted_Cor_max =[np.asarray(x_min_cor)+np.asarray(x_canvas)-1800,np.asarray(x_max_cor)+np.asarray(x_canvas)+1800]
    # Y_adjusted_Cor_min,Y_adjusted_Cor_max = [np.asarray(y_min_cor)+np.asarray(y_canvas)-1500,np.asarray(y_max_cor)+np.asarray(y_canvas)+1500]

    X_adjusted_Cor_min,X_adjusted_Cor_max =[np.asarray(x_min_cor)+np.asarray(x_canvas)-1200,np.asarray(x_max_cor)+np.asarray(x_canvas)+1200]
    Y_adjusted_Cor_min,Y_adjusted_Cor_max = [np.asarray(y_min_cor)+np.asarray(y_canvas)-1500,np.asarray(y_max_cor)+np.asarray(y_canvas)+1000]
    
    Width_cor_list,Height_cor_list = [X_adjusted_Cor_max- X_adjusted_Cor_min,Y_adjusted_Cor_max- Y_adjusted_Cor_min]
    X_cor_list,Y_cor_list = [X_adjusted_Cor_min,Y_adjusted_Cor_min]
    Filelist,folder_list = [File_crop_list,Folder_crop_list]
    
    All_cor = [Filelist,X_cor_list,Y_cor_list,Width_cor_list,Height_cor_list,Folder_crop_list,Tray_crop_list]
    All_cor= np.array([All_cor]).T.tolist()      


    
    t0 = time.time()
    for filename,X_cor,Y_cor,Width_cor,Height_cor,folder,Tray in All_cor:
        #step1 crop image from SVS file 
        DIRECTORY = r'C:\Users\WUPEIZH\Desktop\temp'
        DIRECTORY = os.path.join(DIRECTORY,'Tray'+str(Tray[0]))# save segmented SVS file into here
        if not os.path.exists(DIRECTORY):
            os.makedirs(DIRECTORY)
        print('working on step1: crop image from SVS file ')
        save_to_path = os.path.join(DIRECTORY,filename[0],'Image')# save segmented SVS file into here
        if os.path.exists(save_to_path):
            print('Background already exist for file '+filename[0])
            continue
        if not os.path.exists(save_to_path):
            os.makedirs(save_to_path)
        filename = filename[0]
        full = os.path.join(folder[0],filename)
        read_slide(full,xstart = int(X_cor[0]),ystart =int(Y_cor[0]),width = int(Width_cor[0]),height = int(Height_cor[0]))
        
        find_crop_coordinate(read_slide.image)    
        NumoftileX = np.arange(0,find_crop_coordinate.NX,1)
        NumoftileY = np.arange(0,find_crop_coordinate.NY,1)
        round = 0
        for x in NumoftileX:
            xstart = 0+2400*x
            for y in NumoftileY:
                ystart = 0+3600*y
                print(xstart,ystart)
                round += 1
                crop_image(read_slide.image,cropxstart = xstart,cropystart = ystart,roundn = round,name = filename,path = save_to_path )
    
    
        #step2 use AI model to generate mask    
        Filelist = os.listdir(DIRECTORY)
        for filename in Filelist:
            read_from_path = os.path.join(DIRECTORY,filename,'Image')# save segmented SVS file into here 
            if len(os. listdir(read_from_path)) == 0:
                print('No Background image in '+read_from_path)
                continue
            save_to_path = os.path.join(DIRECTORY,filename,'Mask')# save segmented SVS file into here 
            if os.path.exists(save_to_path):
                print('Mask already exist for file '+filename)
                continue
            if not os.path.exists(save_to_path):
                os.makedirs(save_to_path)
            DATA_DIR = read_from_path # r'C:\Users\WUPEIZH\Desktop\temp'
            print('working on step2: generate masks')
            x_test_dir = DATA_DIR
            y_test_dir = DATA_DIR
            test_dataset = Dataset(
                x_test_dir, 
                y_test_dir, 
                augmentation=get_validation_augmentation(), 
                preprocessing=get_preprocessing(preprocessing_fn),
                classes=CLASSES,
                )
            test_dataloader = DataLoader(test_dataset)
            allfiles= os.listdir(read_from_path)
            for i in range(0,len(allfiles)):
                n=i
                image_vis = test_dataset[n][0].astype('uint8')
                image_vis = image_vis.transpose(1,2,0).astype('uint8')
                image, gt_mask = test_dataset[n]
                gt_mask = gt_mask.squeeze()
                x_tensor = torch.from_numpy(image).to(DEVICE).unsqueeze(0)
                pr_mask_red = best_model_red.predict(x_tensor)
                pr_mask_blue = best_model_blue.predict(x_tensor)
                pr_mask_red_list = (pr_mask_red.squeeze().cpu().numpy().round())
                pr_mask_blue_list = (pr_mask_blue.squeeze().cpu().numpy().round())
                pr_mask_list = pr_mask_red_list +pr_mask_blue_list
                pr_mask_list[pr_mask_list ==2]=1
                pr_mask_uint8 = pr_mask_list.astype(np.uint8)
                #pre_mask.append( pr_mask)
                name = test_dataset.ids[n]
                cv2.imwrite(os.path.join(save_to_path,name),pr_mask_uint8)
                print('mask writen to ' + os.path.join(save_to_path,name))
            background_path = os.path.join(DIRECTORY,filename,'Image')# save segmented SVS file into here 
            mask_path = os.path.join(DIRECTORY,filename,'Mask')# save segmented SVS file into here
            if os.path.exists(mask_path) ==0 or os.path.exists(background_path) == 0:
                print('No Mask image or background image in '+filename)
                continue
            
            
        # step 3 make overlay and store data
            print('working on step3: overlay and stats')
            overlay_path = os.path.join(DIRECTORY,filename,'Overlay')
            if os.path.exists(overlay_path):
                print('Overlay already exist for file '+filename)
                continue
            if not os.path.exists(overlay_path):
                os.makedirs(overlay_path)
            allfiles= os.listdir(background_path)
            segment_name = ['Total counts for ' + filename]
            segment_stat = [[0, 0, 0, 0, 0]]
            centroid_stat = [[0, 0]]
            for name in allfiles:
                backgroun_path_seg = os.path.join(background_path,name)
                mask_path_seg = os.path.join(mask_path,name)
                overlay_path_seg = os.path.join(overlay_path,name)
                if os.path.isfile(backgroun_path_seg) == 0 or os.path.isfile(mask_path_seg) ==0:
                    print('No Mask image or background image in '+ filename + ' ' + name)
                    continue
                background = cv2.imread(backgroun_path_seg)
                #background = cv2.cvtColor(background,cv2.COLOR_BGR2RGB)
                mask = cv2.imread(mask_path_seg)
                mask_modified = mask[0:2400,7:3607,0:3]*255
                mask_modified[:,:,0]=0
                mask_modified[:,:,2]=0
                overlay = cv2.addWeighted(background,1,mask_modified,1,1)
                cv2.imwrite(overlay_path_seg,overlay)
                # the following two lines are to count the number of incidence 
                gray = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
                retval, labels, stats, centroids = cv2.connectedComponentsWithStats(gray)
                segment_name = segment_name+[name]*len(stats)
                segment_stat=np.concatenate([segment_stat,stats],axis = 0)
                centroid_stat=np.concatenate([centroid_stat,centroids],axis = 0)
                NumberofSGN_in_segment = sum((stats[:,4]>0) & (stats[:,4]<250))
                if NumberofSGN_in_segment == 0 :
                    os.remove(backgroun_path_seg)
                    os.remove(mask_path_seg)
                    os.remove(overlay_path_seg)
                    print('no SGN, remove '+ filename, name)
            # save the stats to excel
            segment_stat[0,0] = sum((segment_stat[:,4]>0) & (segment_stat[:,4]<250))
            segment_name = np.array([segment_name]).T.tolist()
            save_to_file = pd.DataFrame(np.concatenate([segment_name,segment_stat,centroid_stat],axis = 1),
                                            columns = ['name',
                                                        'leftmost cor',
                                                        'topmost cor',
                                                        'horizontal size',
                                                        'vertical size',
                                                        'total area',
                                                        'X centroids',
                                                        'Y centroids'])
            excel_path = os.path.join(DIRECTORY,filename,filename+'stats.xlsx')
            save_to_file.to_excel(excel_path)  
            print('Data exported excel done for '+filename)
    
    
        
        
    t1 = time.time()
    print ('total runtime = ',(t1-t0)/60,'mins')


    #%% create merged image and count the number of masks within an image and save the stat.excel
    DIRECTORY = r'C:\Users\WUPEIZH\Desktop\temp'
    tray = 'Tray'+Tray[0]
    Slides = os.listdir(os.path.join(DIRECTORY,tray))
    overlay_path =os.path.join(savetopath,'Merged')
    if not os.path.exists(overlay_path):
        os.makedirs(overlay_path)
    Mask_path =os.path.join(savetopath,'Mask')
    if not os.path.exists(Mask_path):
        os.makedirs(Mask_path)
    Stat_path =os.path.join(savetopath,'stat')
    if not os.path.exists(Stat_path):
        os.makedirs(Stat_path)
        
    for slide in Slides: 
        Slide_stat = [[0, 0, 0, 0, 0]]
        centroid_stat = [[0, 0]]
        IndexN = Filelist.index(slide)
        width,height = [Width_cor_list[IndexN],Height_cor_list[IndexN]]
        NumOfSegPerRow,NumOfSegPerCol = [math.ceil(width/2400),math.ceil(height/3600)]
        read_slide(os.path.join(folder_list[IndexN],Filelist[IndexN]),X_cor_list[IndexN],Y_cor_list[IndexN],width,height)
        MergedMask = np.zeros((NumOfSegPerCol*3600,NumOfSegPerRow*2400,3))
        Background = MergedMask.astype((np.uint8))
        Background[0:height,0:width,0:3]= read_slide.image
        Background = cv2.cvtColor(Background,cv2.COLOR_BGR2RGB)
        MergedMask = MergedMask.astype((np.uint8))
        name = os.path.join(DIRECTORY,tray,slide)
        Mergeoverlay = os.path.join(overlay_path,Filelist[IndexN]+'.merge_overlay.tif')
        Mergemask = os.path.join(Mask_path,Filelist[IndexN]+'.merge_mask.tif')
        Original,Masks,Overlay = [os.path.join(name,'Image'), os.path.join(name,'Mask'), os.path.join(name,'Overlay')]
        Segments = os.listdir(Original)   
        for eachsegment in Segments:
            result = re.search('segment(.*).tif', eachsegment)
            segmentnumber = int(result.group(1))
            PositionOfColumn = math.ceil(segmentnumber/NumOfSegPerCol)
            PositionOfRow = segmentnumber-(PositionOfColumn-1)*NumOfSegPerCol
            if os.path.isfile(os.path.join(Masks,eachsegment)):
                slide1 = cv2.imread(os.path.join(Masks,eachsegment))
                mask_modified = slide1[0:2400,7:3607,0:3]*255
                mask_modified[:,:,0]=0
                mask_modified[:,:,2]=0
                New =np.flip( np.rot90(mask_modified))
                MergedMask[3600*(PositionOfRow-1):3600*(PositionOfRow-1)+3600,2400*(PositionOfColumn-1):2400*(PositionOfColumn-1)+2400,0:3]=New
        overlay = cv2.addWeighted(Background,1,MergedMask,1,1)
        overlay = overlay[0:height,0:width,0:3]
        cv2.imwrite(Mergeoverlay,overlay)    
        cv2.imwrite(Mergemask,MergedMask)  
        gray = cv2.cvtColor(MergedMask, cv2.COLOR_BGR2GRAY)
        retval, labels, stats, centroids = cv2.connectedComponentsWithStats(gray)
        Slide_stat=np.concatenate([Slide_stat,stats],axis = 0)
        centroid_stat=np.concatenate([centroid_stat,centroids],axis = 0)
        Slide_stat[0,0] = sum((stats[:,4]>0) & (stats[:,4]<250))
        save_to_file = pd.DataFrame(np.concatenate([Slide_stat,centroid_stat],axis = 1),
                                    columns = ['leftmost cor','topmost cor',
                                                'horizontal size','vertical size',
                                                'total area','X centroids','Y centroids'])
        excel_path = os.path.join(Stat_path,slide+'stats.xlsx')
        save_to_file.to_excel(excel_path)  
    
        
    #%% get all the SGN location (read from excel)    
    ReadfromDatasource = Datasource.active
    
    Stat_path =os.path.join(savetopath,'stat')
    Allfiles = os.listdir(Stat_path)
    x_cor_largecanvas,y_cor_largecanvas,megaFilelist = [[],[],[]]
    for row in range (first_line,last_line):
        readexcel(ReadfromDatasource,row)
        x_cor_largecanvas,y_cor_largecanvas = [x_cor_largecanvas+[int(readexcel.X_cor[0])], y_cor_largecanvas+[int(readexcel.Y_cor[0])]]
        megaFilelist = megaFilelist + [readexcel.Filelist[0]]
    
    x_list_SGN,y_list_SGN,z_list_SGN = [[],[],[]]
    for index,file in enumerate(Allfiles):
        Data = load_workbook(os.path.join(Stat_path,file))
        Dataactive = Data.active
        x_SGN,y_SGN = [Readcoordinate(Dataactive,'G',[],1),Readcoordinate(Dataactive,'H',[],1)]
        Sli = re.search(Side+'(.*).svs',file)
        if 'r' in Sli.group(1):
            index = Sli.group(1).index('r') #number+'r' is rescanned image
            numberofSGNimage = int(Sli.group(1)[0:index])
        else:
            numberofSGNimage = int(Sli.group(1))
            
        z_SGN = [numberofSGNimage]*len(x_SGN)
        Sli = re.search('-(.*).svs',file)   
        for ind,megafile in enumerate(megaFilelist):
            if Sli.group(1) in megafile:
                Correct_X_larg,Correct_Y_larg = [x_cor_largecanvas[ind], y_cor_largecanvas[ind]]
        Corrected_x_SGN ,Corrected_y_SGN = [x_SGN+X_adjusted_Cor_min[index]-Correct_X_larg, y_SGN+Y_adjusted_Cor_min[index]-Correct_Y_larg]
        Corrected_x_SGN,Corrected_y_SGN = [Corrected_x_SGN.tolist(),Corrected_y_SGN.tolist()]
        x_list_SGN,y_list_SGN,z_list_SGN = [x_list_SGN+Corrected_x_SGN[2:],
                                            y_list_SGN+Corrected_y_SGN[2:],
                                            z_list_SGN+z_SGN[2:]]    
        
        
    
    #%% rescale SGN to mm save stats
    x_SGN_tofile,y_SGN_tofile,z_SGN_tofile= [savecor(x_list_SGN,[],scale), savecor(y_list_SGN,[],scale),savecor(z_list_SGN,[])]
    x_SGN_toplot,y_SGN_toplot,z_SGN_toplot= [rescale(x_list_SGN,scale),rescale(y_list_SGN,scale),rescale(z_list_SGN,2/100,firstslide)]  
    x_SGN_toplot,y_SGN_toplot,z_SGN_toplot = [x_SGN_toplot.tolist(),y_SGN_toplot.tolist(),z_SGN_toplot.tolist()]
    To_list =np.array( [x_list_SGN,y_list_SGN,z_list_SGN,x_SGN_toplot,y_SGN_toplot,z_SGN_toplot])
    save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','x_SGN', 'y_SGN','z_SGN'])
    save_to_file.to_excel(os.path.join(savetopath,'xyz coordinates for SGN.xlsx'))  
      
    RCDatasource = load_workbook(os.path.join(savetopath,'xyz coordinates for RC.xlsx'))
    ReadfromDatasource = RCDatasource.active
    x_RC_toplot,y_RC_toplot,z_RC_toplot,RC_percentfrombase = [Readcoordinate(ReadfromDatasource,'E',[],1), 
                                        Readcoordinate(ReadfromDatasource,'F',[],1), 
                                        Readcoordinate(ReadfromDatasource,'G',[],1),
                                        Readcoordinate(ReadfromDatasource,'H',[],1)]
    RCCoordinate = (x_RC_toplot,y_RC_toplot,z_RC_toplot,RC_percentfrombase)
    
    SGNDatasource = load_workbook(os.path.join(savetopath,'xyz coordinates for SGN.xlsx'))
    ReadfromDatasource = SGNDatasource.active
    x_SGN_toplot,y_SGN_toplot,z_SGN_toplot = [Readcoordinate(ReadfromDatasource,'E',[],1), 
                                        Readcoordinate(ReadfromDatasource,'F',[],1), 
                                        Readcoordinate(ReadfromDatasource,'G',[],1)]
    SGNCoordinate = (x_SGN_toplot,y_SGN_toplot,z_SGN_toplot)
    
    PercenD_SGN = determinefrequencyforSGN(RCCoordinate,SGNCoordinate)
    To_list =np.array( [x_list_SGN,y_list_SGN,z_list_SGN,x_SGN_toplot,y_SGN_toplot,z_SGN_toplot,PercenD_SGN])
    save_to_file = pd.DataFrame(To_list.T, columns =['xcor','ycor','slide','x_SGN', 'y_SGN','z_SGN','percent from Base'])
    save_to_file.to_excel(os.path.join(savetopath,'xyz coordinates for SGN.xlsx'))  
           
            
            
        
        
        
    
    
    
    
    
    
    
    
    
    